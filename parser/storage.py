"""Excel storage helpers — load previous listings and save new ones."""
import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from rich.console import Console
from . import config

_console = Console()

save_dir: str = config["directories"]["save"]
os.makedirs(save_dir, exist_ok=True)


def get_latest_file():
    """Get the latest Excel file from the save directory."""
    files = [f for f in os.listdir(save_dir) if f.endswith(".xlsx")]
    if not files:
        return None
    files.sort(key=lambda x: os.path.getmtime(os.path.join(save_dir, x)), reverse=True)
    return os.path.join(save_dir, files[0])


def load_previous_data():
    """Load link set from all saved Excel files for deduplication."""
    files = sorted(
        [f for f in os.listdir(save_dir) if f.endswith(".xlsx")],
        key=lambda f: os.path.getmtime(os.path.join(save_dir, f)),
        reverse=True,
    )
    if not files:
        _console.print("[dim]No previous data found.[/dim]")
        return set()
    previous_links: set[str] = set()
    for filename in files:
        filepath = os.path.join(save_dir, filename)
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            assert ws is not None
            for row in ws.iter_rows(min_row=2, values_only=True):
                cell = row[1]
                if isinstance(cell, str):
                    previous_links.add(cell)
            wb.close()
        except (InvalidFileException, AssertionError, KeyError) as e:
            _console.print(f"[yellow]⚠ Skipped {filename}: {e}[/yellow]")
    _console.print(
        f"[dim]Loaded {len(previous_links)} previous listings "
        f"from {len(files)} file(s)[/dim]"
    )
    return previous_links


def load_full_data():
    """Load all listings (price + link) from all saved Excel files."""
    files = sorted(
        [f for f in os.listdir(save_dir) if f.endswith(".xlsx")],
        key=lambda f: os.path.getmtime(os.path.join(save_dir, f)),
        reverse=True,
    )
    all_data = []
    seen: set[str] = set()
    for filename in files:
        filepath = os.path.join(save_dir, filename)
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            assert ws is not None
            for row in ws.iter_rows(min_row=2, values_only=True):
                price, link = row[0], row[1]
                if isinstance(link, str) and link not in seen:
                    seen.add(link)
                    all_data.append({"price": str(price) if price else None, "link": link})
            wb.close()
        except (InvalidFileException, AssertionError, KeyError) as e:
            _console.print(f"[yellow]⚠ Skipped {filename}: {e}[/yellow]")
    return all_data


def clean_data():
    """Delete all saved Excel files."""
    files = [f for f in os.listdir(save_dir) if f.endswith(".xlsx")]
    for f in files:
        os.remove(os.path.join(save_dir, f))
    return len(files)


def save_to_excel(data):
    """Save the collected data to an Excel file."""
    if not data:
        return None
    date_str = datetime.now().strftime("%d %B %H-%M")
    file_path = os.path.join(save_dir, f"njuskalo_listings {date_str}.xlsx")

    def extract_price(entry):
        price = entry["price"]
        if price is None:
            return float("inf")
        match = re.search(r"\d+", price.replace(".", "").replace(",", ""))
        return int(match.group()) if match else float("inf")

    data.sort(key=extract_price)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Price", "Link"])
    link_style = Font(color="0563C1", underline="single")
    for item in data:
        ws.append([item["price"], item["link"]])
        cell = ws.cell(row=ws.max_row, column=2)
        cell.hyperlink = item["link"]
        cell.font = link_style
    wb.save(file_path)
    return file_path
